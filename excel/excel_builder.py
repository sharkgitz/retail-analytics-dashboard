"""
Generates retail_analysis.xlsx from the cleaned retail dataset.
Run after notebooks/eda.ipynb has been executed (it saves data/cleaned/retail_cleaned.csv).

Usage:
    python excel/excel_builder.py

Output:
    excel/retail_analysis.xlsx  (4 sheets with conditional formatting and charts)
"""

import sys
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# --- Paths ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_PATH = os.path.join(PROJECT_ROOT, 'data', 'cleaned', 'retail_cleaned.csv')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'retail_analysis.xlsx')

# --- Colour palette ---
HEADER_BLUE = '1F4E79'
ACCENT_BLUE = 'BDD7EE'
WARN_RED_FILL = 'FFCCCC'
WARN_RED_FONT = 'C00000'
ALT_ROW = 'F2F2F2'
GREEN_FILL = 'E2EFDA'
WHITE = 'FFFFFF'


def load_data() -> pd.DataFrame:
    if not os.path.exists(DATA_PATH):
        print(f'\nERROR: {DATA_PATH} not found.')
        print('Run notebooks/eda.ipynb first to generate the cleaned dataset.\n')
        sys.exit(1)

    df = pd.read_csv(DATA_PATH, parse_dates=['InvoiceDate'])
    if 'Revenue' not in df.columns:
        df['Revenue'] = df['Quantity'] * df['Price']
    if 'YearMonth' not in df.columns:
        df['YearMonth'] = df['InvoiceDate'].dt.to_period('M').astype(str)
    return df


def apply_header_style(ws, row_num: int = 1, fill_hex: str = HEADER_BLUE):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    bold_white = Font(color=WHITE, bold=True, size=10)
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = bold_white
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def auto_column_widths(ws, min_width: int = 10, max_width: int = 40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=min_width
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def thin_border():
    thin = Side(style='thin', color='D9D9D9')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def zebra_rows(ws, start_row: int, end_row: int, num_cols: int):
    """Apply alternating row shading for readability."""
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type='solid')
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = alt_fill


# -----------------------------------------------------------------------
# Sheet 1: Cleaned Data Sample (first 5,000 rows)
# -----------------------------------------------------------------------
def build_raw_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = 'Cleaned Data (5k rows)'

    sample = df.head(5000).copy()
    # Format InvoiceDate as string so Excel doesn't mangle it
    if 'InvoiceDate' in sample.columns:
        sample['InvoiceDate'] = sample['InvoiceDate'].astype(str)

    for r_idx, row in enumerate(dataframe_to_rows(sample, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border()

    apply_header_style(ws, row_num=1)
    ws.row_dimensions[1].height = 22
    auto_column_widths(ws)
    ws.freeze_panes = 'A2'

    # Light zebra shading on data rows
    zebra_rows(ws, start_row=2, end_row=5001, num_cols=len(sample.columns))

    print(f'  Sheet 1: {len(sample):,} rows written')


# -----------------------------------------------------------------------
# Sheet 2: Revenue by Country × Month (pivot-style)
# -----------------------------------------------------------------------
def build_country_month_pivot(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Revenue by Country-Month')

    sales = df[df['Quantity'] > 0].copy()

    # Top 15 countries by total revenue
    top_countries = (
        sales.groupby('Country')['Revenue'].sum()
        .nlargest(15)
        .index.tolist()
    )
    sales = sales[sales['Country'].isin(top_countries)]

    pivot = (
        sales.groupby(['Country', 'YearMonth'])['Revenue']
        .sum()
        .round(2)
        .unstack(level='YearMonth', fill_value=0)
        .reset_index()
    )

    months = [c for c in pivot.columns if c != 'Country']

    # Write header
    ws.cell(row=1, column=1, value='Country')
    for c_idx, month in enumerate(months, 2):
        ws.cell(row=1, column=c_idx, value=month)

    # Write data
    for r_idx, row_data in enumerate(pivot.itertuples(index=False), 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border()
            if c_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    apply_header_style(ws, row_num=1)
    ws.row_dimensions[1].height = 30
    auto_column_widths(ws, min_width=10, max_width=16)
    ws.freeze_panes = 'B2'

    # Conditional formatting: flag months where revenue dropped >10% MoM
    red_fill = PatternFill(start_color=WARN_RED_FILL, end_color=WARN_RED_FILL, fill_type='solid')
    red_font = Font(color=WARN_RED_FONT, bold=True)
    n_rows = len(pivot) + 1  # +1 for header

    for month_idx in range(1, len(months)):
        prev_col = month_idx + 1   # column index of previous month (1-indexed, col 1 = Country)
        curr_col = month_idx + 2

        for row_idx in range(2, n_rows + 1):
            prev_val = ws.cell(row=row_idx, column=prev_col).value or 0
            curr_val = ws.cell(row=row_idx, column=curr_col).value or 0
            if prev_val > 0 and isinstance(curr_val, (int, float)):
                mom_change = (curr_val - prev_val) / prev_val
                if mom_change < -0.10:
                    ws.cell(row=row_idx, column=curr_col).fill = red_fill
                    ws.cell(row=row_idx, column=curr_col).font = red_font

    ws.cell(
        row=n_rows + 2, column=1,
        value='Red cells = revenue dropped >10% vs. prior month'
    ).font = Font(italic=True, color='808080', size=9)

    print(f'  Sheet 2: {len(top_countries)} countries × {len(months)} months')


# -----------------------------------------------------------------------
# Sheet 3: Top 20 Products by Revenue
# -----------------------------------------------------------------------
def build_top_products_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Top 20 Products')

    top20 = (
        df[df['Quantity'] > 0]
        .groupby(['StockCode', 'Description'])
        .agg(
            Total_Revenue=('Revenue', 'sum'),
            Units_Sold=('Quantity', 'sum'),
            Num_Invoices=('Invoice', 'nunique')
        )
        .round(2)
        .nlargest(20, 'Total_Revenue')
        .reset_index()
    )
    total_rev = top20['Total_Revenue'].sum()
    top20['Revenue_Share_Pct'] = (top20['Total_Revenue'] / total_rev * 100).round(2)

    for r_idx, row in enumerate(dataframe_to_rows(top20, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border()

    apply_header_style(ws, row_num=1)
    auto_column_widths(ws)
    ws.freeze_panes = 'A2'

    # Green top-5 highlight
    green_top5 = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type='solid')
    for r in range(2, 7):
        for c in range(1, len(top20.columns) + 1):
            ws.cell(row=r, column=c).fill = green_top5

    zebra_rows(ws, start_row=2, end_row=21, num_cols=len(top20.columns))

    # Horizontal bar chart — revenue by product
    chart = BarChart()
    chart.type = 'bar'
    chart.grouping = 'clustered'
    chart.title = 'Top 20 Products by Gross Revenue (£)'
    chart.y_axis.title = 'Revenue (£)'
    chart.y_axis.numFmt = '#,##0'
    chart.shape = 4
    chart.width = 20
    chart.height = 14

    n_products = len(top20)
    data_ref = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=n_products + 1)
    cats_ref = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=n_products + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, 'H2')

    print(f'  Sheet 3: Top 20 products + bar chart')


# -----------------------------------------------------------------------
# Sheet 4: Returns Analysis with chart
# -----------------------------------------------------------------------
def build_returns_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Returns Analysis')

    monthly = (
        df.groupby('YearMonth')
        .apply(lambda g: pd.Series({
            'Gross_Sales': g.loc[g['Revenue'] > 0, 'Revenue'].sum(),
            'Return_Value': g.loc[g['Revenue'] < 0, 'Revenue'].abs().sum(),
            'Sale_Lines': (g['Quantity'] > 0).sum(),
            'Return_Lines': (g['Quantity'] < 0).sum(),
        }))
        .round(2)
        .reset_index()
    )
    monthly['Net_Sales'] = (monthly['Gross_Sales'] - monthly['Return_Value']).round(2)
    monthly['Return_Rate_Pct'] = (
        np.where(
            monthly['Gross_Sales'] > 0,
            (monthly['Return_Value'] / monthly['Gross_Sales'] * 100).round(2),
            0
        )
    )
    monthly = monthly.sort_values('YearMonth')

    for r_idx, row in enumerate(dataframe_to_rows(monthly, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border()
            if c_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    apply_header_style(ws, row_num=1)
    auto_column_widths(ws)
    ws.freeze_panes = 'A2'
    zebra_rows(ws, start_row=2, end_row=len(monthly) + 1, num_cols=len(monthly.columns))

    # Highlight months where return rate > 8%
    orange_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
    return_rate_col = monthly.columns.get_loc('Return_Rate_Pct') + 1
    for r_idx in range(2, len(monthly) + 2):
        val = ws.cell(row=r_idx, column=return_rate_col).value
        if val is not None and val > 8.0:
            ws.cell(row=r_idx, column=return_rate_col).fill = orange_fill

    # Line chart — return rate over time
    chart = LineChart()
    chart.title = 'Monthly Return Rate (%)'
    chart.y_axis.title = 'Return Rate %'
    chart.y_axis.numFmt = '0.00'
    chart.x_axis.title = 'Month'
    chart.width = 20
    chart.height = 10

    n_months = len(monthly)
    data_ref = Reference(ws, min_col=return_rate_col, max_col=return_rate_col,
                         min_row=1, max_row=n_months + 1)
    cats_ref = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=n_months + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = '1F4E79'
    ws.add_chart(chart, f'I2')

    print(f'  Sheet 4: {len(monthly)} months of returns data + line chart')


# -----------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------
def main():
    print(f'Loading cleaned data from:\n  {DATA_PATH}\n')
    df = load_data()
    print(f'Loaded {len(df):,} rows, {df["Invoice"].nunique():,} unique invoices\n')

    wb = Workbook()

    print('Building sheets...')
    build_raw_sheet(wb, df)
    build_country_month_pivot(wb, df)
    build_top_products_sheet(wb, df)
    build_returns_sheet(wb, df)

    wb.save(OUTPUT_PATH)
    print(f'\nSaved: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
